from bs4 import BeautifulSoup
import requests
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
import datetime
from time import sleep
from selenium import webdriver

################################################################################
#############################Variables##########################################
################################################################################

############################Constants###########################################
PAGE_EXTENSION = 10
OFFSET = 18
############################Excel File Variables################################
wb = Workbook()
ws = []
Links_List = []
########################### URL Variables ######################################

EE_first_page = "https://ca.indeed.com/Electrical-Engineer-jobs"
EE_base_page = "https://ca.indeed.com/jobs?q=Electrical+Engineer&start="

SE_first_page = "https://ca.indeed.com/Software-Engineer-jobs"
SE_base_page = "https://ca.indeed.com/jobs?q=Software+Engineer&start="



def get_description(page_url,iteration,ws):
    j = iteration
    chrome = webdriver.Chrome()#open up the chrome web driver
    chrome.get(page_url)#open up the chrome page with the specified url
    chrome.maximize_window()#maximize the chrome window
    sleep(5)
    for i, job_postings in enumerate(chrome.find_elements_by_class_name("title")):

        get_a = job_postings.find_element_by_tag_name("a") #grab the element with tag name "a"
        get_title = get_a.get_attribute("title") # grab the innerhtml within the a tag

        job_postings.find_element_by_partial_link_text(get_title).click()
        sleep(5)

        text = chrome.find_element_by_id("vjs-desc").get_attribute("innerHTML")
        sleep(5)
        cleaned_text = text.replace("<p>","").replace("<b>","").replace("</b>","").replace("</p>","\n").replace("<ul>","").replace("</ul>","\n").replace("<li>","").replace("</li>","\n")
        print(cleaned_text)

        if i == 0:
            ws['F' + str(2 + (OFFSET * j))] = cleaned_text
        else:
            ws['F' + str((2+i) + (OFFSET * j))] = cleaned_text

    sleep(10)
    chrome.quit()#Chrome browser quit

#Programmer: Casey Tse
#Date: September 14,2020
#Parameters: Takes in a worksheet list and an index
#Function: Stylizes the worksheet with the Column Row's and stylizes them to look nice
def Style_Ws(ws):
    ws['A1'] = "Title"
    ws['B1'] = "Company"
    ws['C1'] = "Location"
    ws['D1'] = "Date Posted"
    ws['E1'] = "Indeed Link"
    ws['F1'] = "Job Description"
    ws['A1'].style = "Title"
    ws['B1'].style = "Title"
    ws['C1'].style = "Title"
    ws['D1'].style = "Title"
    ws['E1'].style = "Title"
    ws['F1'].style = "Title"


    ws['A1'].alignment = Alignment(horizontal = "center")
    ws['B1'].alignment = Alignment(horizontal = "center")
    ws['C1'].alignment = Alignment(horizontal = "center")
    ws['D1'].alignment = Alignment(horizontal = "center")
    ws['E1'].alignment = Alignment(horizontal = "center")
    ws['F1'].alignment = Alignment(horizontal = "center")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 85

def Get_all_jobs(first_url,base_url,iterations,File_Ext):
    for i in range(0,iterations):

        if i == 0:
            ws = wb.active
            ws.title = "Page " + str(i) #Title means its the title of the page itself.
            Style_Ws(ws)
            get_job(first_url,i,ws)

        else:
            get_job(base_url + str(PAGE_EXTENSION*i),i,ws)

    wb.save(filename = Get_file_name(File_Ext))

#Programmer: Casey Tse
#Date: September 15,2020
#Parameters: Asks for the File Extension
#Function: Returns a string + current date to be used in saving Excel File Name
def Get_file_name(File_Ext):
    x = datetime.datetime.now()
    year = x.strftime("%Y")
    month = x.strftime("%m")
    day = x.strftime("%d")
    date = (day + "_" + month + "_" + year)
    return ("Job_Postings for_" + File_Ext + "_" + date +".xlsx")

#Programmer: Casey Tse
#Date: September 15,2020
#Parameters: Requires a list full of links, gotten from get_job
#Function: Goes through the list of links, and requests there page,
#Grabs the description of the job, and places it into a excel file
def get_job(page_url,iteration,ws):
    j = iteration
    get_description(page_url,j,ws)
    sleep(15)
    page = requests.get(page_url)
    if not page.status_code == 200:
        print("Page Request Failed!")
    else:
        soup = BeautifulSoup(page.content,'html.parser')
        for i,postings in enumerate(soup.find_all(class_="jobsearch-SerpJobCard")):

            job_title = postings.find(class_="jobtitle turnstileLink")
            Title_Posting = job_title.get("title")
            link = job_title.get("href")
            Indeed_Link = "https://ca.indeed.com/" + link

            date = postings.find(class_="date")
            Date_Posted = date.get_text()

            Company = postings.find(class_="company")
            Company_Posting = Company.get_text()

            Location = postings.find(class_="location accessible-contrast-color-location")
            Location_Posting = Location.get_text()

            if i == 0:
                ws['A' + str(2 +(OFFSET * j))] = Title_Posting
                ws['B' + str(2 +(OFFSET * j))] = Company_Posting
                ws['C' + str(2 +(OFFSET * j))] = Location_Posting
                ws['D' + str(2 +(OFFSET * j))] = Date_Posted
                ws['E' + str(2 +(OFFSET * j))] = Indeed_Link
                ws['E' + str(2 +(OFFSET * j))].style = "Hyperlink"
            else:
                ws['A' + str((2 + i) + (OFFSET * j))] = Title_Posting
                ws['B' + str((2 + i) + (OFFSET * j))] = Company_Posting
                ws['C' + str((2 + i) + (OFFSET * j))] = Location_Posting
                ws['D' + str((2 + i) + (OFFSET * j))] = Date_Posted
                ws['E' + str((2 + i) + (OFFSET * j))] = Indeed_Link
                ws['E' + str((2 + i) + (OFFSET * j))].style = "Hyperlink"
                ws['E' + str((2 + i) + (OFFSET * j))].alignment = Alignment(horizontal = "fill")
        print("Job_Postings saved to Excel File!")

        #text allignment = fill

if __name__ == "__main__":
    print("Hello World!")
    # generate_array(ws,1)
    Get_all_jobs(EE_first_page,EE_base_page,1,"EE")
    # Get_all_jobs(SE_first_page,SE_base_page,1,"SE")
